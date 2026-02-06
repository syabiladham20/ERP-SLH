import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import TextAxis
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_template():
    wb = Workbook()

    # --- Sheet 1: Flock Reference ---
    ws_ref = wb.active
    ws_ref.title = "Flock Reference"

    headers_ref = ["Flock ID", "Intake Date"]
    ws_ref.append(headers_ref)

    # Styling Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws_ref[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws_ref["A2"] = "Example_Flock_01"
    ws_ref["B2"] = "2023-01-01"
    ws_ref.column_dimensions["A"].width = 25
    ws_ref.column_dimensions["B"].width = 15

    # --- Sheet 2: Data ---
    ws_data = wb.create_sheet("Data")

    headers_data = [
        "Setting Date",       # A
        "Candling Date",      # B
        "Hatching Date",      # C
        "Flock ID",           # D
        "Egg Set",            # E
        "Clear Egg",          # F
        "Clear Egg %",        # G (Calc)
        "Rotten Egg",         # H
        "Rotten Egg %",       # I (Calc)
        "Hatchable Egg",      # J (Calc)
        "Hatchable Egg %",    # K (Calc)
        "Total Hatched",      # L
        "Hatchability %",     # M (Calc)
        "Male Ratio %",       # N (Input)
        "Flock Age (Weeks)"   # O (Calc)
    ]

    ws_data.append(headers_data)

    for cell in ws_data[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column Widths
    ws_data.column_dimensions['A'].width = 12
    ws_data.column_dimensions['B'].width = 12
    ws_data.column_dimensions['C'].width = 12
    ws_data.column_dimensions['D'].width = 20
    ws_data.column_dimensions['E'].width = 10
    ws_data.column_dimensions['F'].width = 10
    ws_data.column_dimensions['G'].width = 12
    ws_data.column_dimensions['H'].width = 10
    ws_data.column_dimensions['I'].width = 12
    ws_data.column_dimensions['J'].width = 12
    ws_data.column_dimensions['K'].width = 15
    ws_data.column_dimensions['L'].width = 12
    ws_data.column_dimensions['M'].width = 15
    ws_data.column_dimensions['N'].width = 12
    ws_data.column_dimensions['O'].width = 15

    # Define Data Table for easier referencing in Analysis
    max_row = 1000
    tab = Table(displayName="HatchData", ref=f"A1:O{max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws_data.add_table(tab)

    # Define Formulas (Rows 2 to max_row)
    for row in range(2, 502):
        ws_data[f"G{row}"] = f"=IF(E{row}>0, F{row}/E{row}, 0)" # Clear %
        ws_data[f"I{row}"] = f"=IF(E{row}>0, H{row}/E{row}, 0)" # Rotten %
        ws_data[f"J{row}"] = f"=IF(E{row}>0, E{row}-F{row}-H{row}, 0)" # Hatchable/Fertile
        ws_data[f"K{row}"] = f"=IF(E{row}>0, J{row}/E{row}, 0)" # Fertile %
        ws_data[f"M{row}"] = f"=IF(E{row}>0, L{row}/E{row}, 0)" # Hatchability %
        # Age Calculation:
        # Intake = Day 0.
        # Week 0 = Day 1 to Day 7. (Day 1 - 1)/7 = 0. (Day 7 - 1)/7 = 0.
        # Week 1 = Day 8 to Day 14. (Day 8 - 1)/7 = 1.
        # "Use Offset": Use age of the day BEFORE Setting Date.
        # Formula: Age = INT( ( (SettingDate - 1) - IntakeDate - 1 ) / 7 ) = INT( (Setting - Intake - 2) / 7 )
        ws_data[f"O{row}"] = f"=IFERROR(INT((A{row}-2-VLOOKUP(D{row},'Flock Reference'!$A:$B,2,FALSE))/7), \"\")"

        for col in ['G', 'I', 'K', 'M', 'N']:
            ws_data[f"{col}{row}"].number_format = '0.00%'

    # --- Sheet 3: Dashboard / Analysis ---
    ws_dash = wb.create_sheet("Analysis")

    # Inputs
    ws_dash["A1"] = "Select Flock:"
    ws_dash["B1"] = "Example_Flock_01" # Default
    ws_dash["A2"] = "Start Age:"
    ws_dash["B2"] = 0
    ws_dash["A3"] = "End Age:"
    ws_dash["B3"] = 100

    ws_dash["A1"].font = Font(bold=True)
    ws_dash["A2"].font = Font(bold=True)
    ws_dash["A3"].font = Font(bold=True)
    ws_dash["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Highlight Input

    # --- Dynamic Aggregation Table ---
    # We will use Excel Dynamic Array formulas (SORT, UNIQUE, FILTER)
    # Target Structure:
    # Age | Sum Egg Set | Sum Clear | Sum Rotten | Sum Hatched | Fertile | Fertile% | Clear% | Rotten% | Hatch%

    # Header
    agg_headers = ["Age (Weeks)", "Egg Set", "Clear", "Rotten", "Hatched", "Fertile", "Fertile %", "Clear %", "Rotten %", "Hatchability %"]
    ws_dash.append([]) # Spacer
    ws_dash.append([]) # Spacer
    ws_dash.append(agg_headers) # Row 6

    for cell in ws_dash[6]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # A7: Unique Ages based on Filter
    # =SORT(UNIQUE(FILTER(HatchData[Flock Age (Weeks)], (HatchData[Flock ID]=B1)*(HatchData[Flock Age (Weeks)]>=B2)*(HatchData[Flock Age (Weeks)]<=B3))))
    # Note: Column names in Table Ref must match exactly.
    # Col O Name: "Flock Age (Weeks)"
    # Col D Name: "Flock ID"

    formula_unique_ages = """=SORT(UNIQUE(FILTER(HatchData[Flock Age (Weeks)], (HatchData[Flock ID]=B1)*(HatchData[Flock Age (Weeks)]>=B2)*(HatchData[Flock Age (Weeks)]<=B3), "No Data")))"""
    ws_dash["A7"] = formula_unique_ages

    # B7: Sum Egg Set
    # =SUMIFS(HatchData[Egg Set], HatchData[Flock ID], $B$1, HatchData[Flock Age (Weeks)], A7#)
    # using # to reference spill range
    ws_dash["B7"] = """=SUMIFS(HatchData[Egg Set], HatchData[Flock ID], $B$1, HatchData[Flock Age (Weeks)], A7#)"""

    # C7: Sum Clear
    ws_dash["C7"] = """=SUMIFS(HatchData[Clear Egg], HatchData[Flock ID], $B$1, HatchData[Flock Age (Weeks)], A7#)"""

    # D7: Sum Rotten
    ws_dash["D7"] = """=SUMIFS(HatchData[Rotten Egg], HatchData[Flock ID], $B$1, HatchData[Flock Age (Weeks)], A7#)"""

    # E7: Sum Hatched
    ws_dash["E7"] = """=SUMIFS(HatchData[Total Hatched], HatchData[Flock ID], $B$1, HatchData[Flock Age (Weeks)], A7#)"""

    # F7: Calc Fertile (Set - Clear - Rotten) -> Aggregated
    # We can't use SUMIFS easily on calculated column J unless J is robust, but simpler to recalc: B7# - C7# - D7#
    ws_dash["F7"] = """=B7#-C7#-D7#"""

    # G7: Fertile % (F / B)
    ws_dash["G7"] = """=IF(B7#>0, F7#/B7#, 0)"""

    # H7: Clear % (C / B)
    ws_dash["H7"] = """=IF(B7#>0, C7#/B7#, 0)"""

    # I7: Rotten % (D / B)
    ws_dash["I7"] = """=IF(B7#>0, D7#/B7#, 0)"""

    # J7: Hatchability % (E / B)
    ws_dash["J7"] = """=IF(B7#>0, E7#/B7#, 0)"""

    # Formatting
    # We can't pre-format the spill range easily in openpyxl without guessing size,
    # but we can format the first few hundred rows of those columns
    for r in range(7, 107):
        for c in ['G', 'H', 'I', 'J']:
            ws_dash[f"{c}{r}"].number_format = '0.00%'

    # --- Chart ---
    # Chart references must point to the spill range.
    # Openpyxl doesn't support "A7#" reference syntax in charts directly easily.
    # We usually define a Name Manager name or reference a fixed large range.
    # For robust template, let's reference A7:A100 (assuming < 100 weeks).

    chart_max_row = 100

    c1 = BarChart()
    c1.type = "col"
    c1.style = 10
    c1.grouping = "stacked"
    c1.overlap = 100
    c1.title = "Aggregated Hatchability Analysis"
    c1.y_axis.title = "Egg Composition (%)"
    c1.x_axis.title = "Flock Age (Weeks)"

    # X-Axis: A7...
    cats = Reference(ws_dash, min_col=1, min_row=7, max_row=chart_max_row) # Age
    c1.set_categories(cats)

    # Series:
    # Fertile % (G)
    # Clear % (H)
    # Rotten % (I)

    data_fertile = Reference(ws_dash, min_col=7, min_row=7, max_row=chart_max_row)
    data_clear = Reference(ws_dash, min_col=8, min_row=7, max_row=chart_max_row)
    data_rotten = Reference(ws_dash, min_col=9, min_row=7, max_row=chart_max_row)

    s1 = Series(data_fertile, title="Fertile %")
    s2 = Series(data_clear, title="Clear %")
    s3 = Series(data_rotten, title="Rotten %")

    c1.append(s1)
    c1.append(s2)
    c1.append(s3)

    # Line Chart: Hatchability % (J)
    c2 = LineChart()
    c2.style = 13
    c2.y_axis.title = "Hatchability %"
    c2.y_axis.axId = 200
    c2.y_axis.crosses = "max"

    data_hatch = Reference(ws_dash, min_col=10, min_row=7, max_row=chart_max_row)
    s4 = Series(data_hatch, title="Hatchability %")
    c2.append(s4)

    c1 += c2

    c1.height = 15
    c1.width = 30

    ws_dash.add_chart(c1, "L2")

    wb.save("Hatchability_Template.xlsx")
    print("Template created: Hatchability_Template.xlsx")

if __name__ == "__main__":
    create_template()
